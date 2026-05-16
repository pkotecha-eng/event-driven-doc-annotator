import io
import csv
import json
from pathlib import Path


def extract_text(filename: str, content: bytes) -> dict:
    """
    Extract raw text and metadata from uploaded document.
    Supports PDF and spreadsheet formats (csv, xlsx, xls).
    Returns dict with 'text', 'doc_format', 'page_count' (if applicable).
    """
    suffix = Path(filename).suffix.lower()

    if suffix == ".pdf":
        return _extract_pdf(content)
    elif suffix == ".csv":
        return _extract_csv(content)
    elif suffix in (".xlsx", ".xls"):
        return _extract_excel(content)
    else:
        # Best-effort: treat as plain text
        try:
            text = content.decode("utf-8", errors="replace")
            return {"text": text[:10000], "doc_format": "text", "page_count": None}
        except Exception:
            return {"text": "", "doc_format": "unknown", "page_count": None}


def _extract_pdf(content: bytes) -> dict:
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(content))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        full_text = "\n\n".join(pages)
        return {
            "text": full_text[:15000],  # cap to avoid token blowout
            "doc_format": "pdf",
            "page_count": len(reader.pages)
        }
    except Exception as e:
        return {"text": "", "doc_format": "pdf", "page_count": None, "extract_error": str(e)}


def _extract_csv(content: bytes) -> dict:
    try:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        # Represent as readable text for LLM: header + first 50 rows
        if not rows:
            return {"text": "", "doc_format": "csv", "page_count": None}
        header = rows[0]
        preview_rows = rows[1:51]
        lines = [",".join(header)]
        for row in preview_rows:
            lines.append(",".join(row))
        summary = f"CSV with {len(rows)-1} data rows and {len(header)} columns.\n\n"
        summary += "\n".join(lines)
        return {"text": summary[:15000], "doc_format": "csv", "page_count": None, "row_count": len(rows)-1, "col_count": len(header)}
    except Exception as e:
        return {"text": "", "doc_format": "csv", "page_count": None, "extract_error": str(e)}


def _extract_excel(content: bytes) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 50:
                    break
                rows.append(",".join(str(c) if c is not None else "" for c in row))
            all_text.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        full_text = "\n\n".join(all_text)
        return {
            "text": full_text[:15000],
            "doc_format": "xlsx",
            "page_count": None,
            "sheet_count": len(wb.sheetnames)
        }
    except Exception as e:
        return {"text": "", "doc_format": "xlsx", "page_count": None, "extract_error": str(e)}
